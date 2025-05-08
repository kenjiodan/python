import openpyxl
from statistics import mean, median, mode
import os

NOME_PASTA = "dados_excel"
NOME_ARQUIVO_EXCEL = os.path.join(NOME_PASTA, "dados_usuarios.xlsx")

# Código de escape ANSI para texto verde
COR_VERDE = '\033[92m'
COR_PADRAO = '\033[0m'

if not os.path.exists(NOME_PASTA):
    os.makedirs(NOME_PASTA)

def limpar_tela():
    """Limpa a tela do terminal."""
    os.system('cls' if os.name == 'nt' else 'clear')

def adicionar_usuario(dados):
    """Adiciona um novo usuário à estrutura de dados na memória, inicializando acessos e tempo de uso."""
    limpar_tela()
    print(f"{COR_VERDE}--- Adicionar Novo Usuário ---{COR_PADRAO}")
    while True:
        nome = input("Digite o nome do usuário: ")
        if nome:
            break
        else:
            print("O nome do usuário não pode ser vazio.")

    while True:
        try:
            idade = int(input("Digite a idade do usuário: "))
            if idade >= 0:
                break
            else:
                print("A idade deve ser um número inteiro não negativo.")
        except ValueError:
            print("Entrada inválida. Digite um número inteiro para a idade.")

    # Inicializa o número de acessos em 1 (simulando o primeiro login)
    acessos = 1
    print(f"{COR_VERDE}Acessos para '{nome}' inicializado em: {acessos}{COR_PADRAO}")

    # Inicializa o tempo de uso com 0 (seria rastreado dinamicamente em um sistema real)
    tempo_uso = 0.0
    print(f"{COR_VERDE}Tempo de uso para '{nome}' inicializado em: {tempo_uso:.2f} minutos.{COR_PADRAO}")

    dados.append({"nome": nome, "idade": idade, "acessos": acessos, "tempo_uso": tempo_uso})
    print(f"{COR_VERDE}Usuário '{nome}' adicionado com sucesso!{COR_PADRAO}")
    input("Pressione Enter para voltar ao menu...") # Adiciona uma pausa antes de voltar ao menu

def carregar_dados_excel():
    """Carrega os dados existentes do arquivo Excel com tratamento de erros."""
    limpar_tela()
    print(f"{COR_VERDE}--- Carregar Dados Existentes ---{COR_PADRAO}")
    try:
        workbook = openpyxl.load_workbook(NOME_ARQUIVO_EXCEL)
        sheet = workbook.active
        dados = []
        cabecalhos = [cell.value for cell in sheet[1]]
        if cabecalhos != ["nome", "idade", "acessos", "tempo_uso"]:
            print("Aviso: O arquivo Excel não possui o cabeçalho esperado. Carregando os dados como estão.")
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if all(cell is not None for cell in row):
                    dados.append(dict(zip(["coluna1", "coluna2", "coluna3", "coluna4"], row)))
        else:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is not None for cell in row):
                    dados.append(dict(zip(cabecalhos, row)))
        print(f"{COR_VERDE}Dados carregados com sucesso de '{NOME_ARQUIVO_EXCEL}'.{COR_PADRAO}")
        return dados
    except FileNotFoundError:
        print(f"Arquivo '{NOME_ARQUIVO_EXCEL}' não encontrado. Iniciando com uma lista de dados vazia.")
        return []
    except Exception as e:
        print(f"Ocorreu um erro ao carregar o arquivo Excel: {e}")
        return []
    input("Pressione Enter para voltar ao menu...") # Adiciona uma pausa antes de voltar ao menu

def salvar_dados_excel(dados):
    """Salva a estrutura de dados atualizada no arquivo Excel com tratamento de erros."""
    limpar_tela()
    print(f"{COR_VERDE}--- Salvar Dados ---{COR_PADRAO}")
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["nome", "idade", "acessos", "tempo_uso"])
        for usuario in dados:
            sheet.append([usuario["nome"], usuario["idade"], usuario["acessos"], usuario["tempo_uso"]])
        workbook.save(NOME_ARQUIVO_EXCEL)
        print(f"{COR_VERDE}Dados salvos com sucesso em '{NOME_ARQUIVO_EXCEL}'.{COR_PADRAO}")
    except Exception as e:
        print(f"Ocorreu um erro ao salvar o arquivo Excel: {e}")
    input("Pressione Enter para voltar ao menu...") # Adiciona uma pausa antes de voltar ao menu

def calcular_estatisticas(dados):
    """Calcula e exibe a média, moda e mediana para as características numéricas."""
    print(f"{COR_VERDE}--- Calcular e Exibir Estatísticas ---{COR_PADRAO}")
    if not dados:
        print("Não há dados para calcular as estatísticas.")
        return

    idades = [usuario["idade"] for usuario in dados if "idade" in usuario and isinstance(usuario["idade"], (int, float))]
    acessos = [usuario["acessos"] for usuario in dados if "acessos" in usuario and isinstance(usuario["acessos"], (int, float))]
    tempos_uso = [usuario["tempo_uso"] for usuario in dados if "tempo_uso" in usuario and isinstance(usuario["tempo_uso"], (int, float))]

    print(f"{COR_VERDE}--- Estatísticas dos Usuários ---{COR_PADRAO}")

    if idades:
        try:
            media_idade = mean(idades)
            mediana_idade = median(idades)
            moda_idade = mode(idades) if len(set(idades)) < len(idades) else 'Não há moda clara'
            print(f"{COR_VERDE}Idade - Média: {media_idade:.2f}, Mediana: {mediana_idade}, Moda: {moda_idade}{COR_PADRAO}")
            print("Cálculo da Média (Idade): Soma de todas as idades / Número total de usuários.")
            print("Cálculo da Mediana (Idade): Valor central quando as idades são ordenadas.")
            print("Cálculo da Moda (Idade): Valor que aparece com mais frequência.")
        except Exception as e:
            print(f"Erro ao calcular estatísticas de idade: {e}")
    else:
        print("Não há dados de idade para calcular as estatísticas.")

    if acessos:
        try:
            media_acessos = mean(acessos)
            mediana_acessos = median(acessos)
            moda_acessos = mode(acessos) if len(set(acessos)) < len(acessos) else 'Não há moda clara'
            print(f"{COR_VERDE}Acessos - Média: {media_acessos:.2f}, Mediana: {mediana_acessos}, Moda: {moda_acessos}{COR_PADRAO}")
            print("Cálculo da Média (Acessos): Soma de todos os acessos / Número total de usuários.")
            print("Cálculo da Mediana (Acessos): Valor central quando o número de acessos é ordenado.")
            print("Cálculo da Moda (Acessos): Valor que aparece com mais frequência.")
        except Exception as e:
            print(f"Erro ao calcular estatísticas de acessos: {e}")
    else:
        print("Não há dados de acessos para calcular as estatísticas.")

    if tempos_uso:
        try:
            media_tempo_uso = mean(tempos_uso)
            mediana_tempo_uso = median(tempos_uso)
            moda_tempo_uso = mode(tempos_uso) if len(set(tempos_uso)) < len(tempos_uso) else 'Não há moda clara'
            print(f"{COR_VERDE}Tempo de Uso - Média: {media_tempo_uso:.2f}, Mediana: {mediana_tempo_uso}, Moda: {moda_tempo_uso}{COR_PADRAO}")
            print("Cálculo da Média (Tempo de Uso): Soma de todos os tempos de uso / Número total de usuários.")
            print("Cálculo da Mediana (Tempo de Uso): Valor central quando os tempos de uso são ordenados.")
            print("Cálculo da Moda (Tempo de Uso): Valor que aparece com mais frequência.")
        except Exception as e:
            print(f"Erro ao calcular estatísticas de tempo de uso: {e}")
    else:
        print("Não há dados de tempo de uso para calcular as estatísticas.")
    input("Pressione Enter para voltar ao menu...") # Adiciona uma pausa antes de voltar ao menu

def exibir_menu():
    """Exibe o menu de opções para o usuário com tratamento de erros e limpeza de tela."""
    limpar_tela()
    while True:
        print(f"{COR_VERDE}--- Menu ---{COR_PADRAO}")
        print("1. Adicionar novo usuário e salvar")
        print("2. Carregar dados existentes")
        print("3. Calcular e exibir estatísticas")
        print("4. Sair")
        opcao = input("Escolha uma opção: ")
        if opcao in ["1", "2", "3", "4"]:
            return opcao
        else:
            print("Opção inválida. Por favor, digite um número entre 1 e 4.")

def main():
    """Função principal do programa."""
    dados_usuarios = []

    while True:
        opcao = exibir_menu()

        if opcao == "1":
            adicionar_usuario(dados_usuarios)
        elif opcao == "2":
            dados_usuarios = carregar_dados_excel() # Atualiza a lista com os dados carregados
        elif opcao == "3":
            calcular_estatisticas(dados_usuarios)
        elif opcao == "4":
            limpar_tela()
            print(f"{COR_VERDE}Saindo do programa.{COR_PADRAO}")
            break
        if opcao == "1" or opcao == "2": # Salva após adicionar ou carregar
            salvar_dados_excel(dados_usuarios)

if __name__ == "__main__":
    main()